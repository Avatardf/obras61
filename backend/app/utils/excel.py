"""Utilitários para geração e importação de planilhas Excel."""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Fornecedores ──────────────────────────────────────────────────────────────

_FORN_COLS: list[tuple[str, int]] = [
    ("Nome / Razão Social *", 35),
    ("CNPJ", 20),
    ("Categoria", 15),
    ("Contato", 20),
    ("Telefone", 18),
    ("E-mail", 30),
    ("Cidade", 18),
    ("UF", 5),
    ("Avaliação (1-5)", 14),
    ("Observações", 40),
]

_FORN_EXEMPLOS = [
    ["EXEMPLO: Concremix Ltda", "12.345.678/0001-90", "Concreto", "João Silva",
     "(21) 99999-0000", "joao@concremix.com.br", "Rio de Janeiro", "RJ", 4,
     "Fornecedor preferencial"],
    ["EXEMPLO: Madeireira Porto", "98.765.432/0001-01", "Madeira", "",
     "(11) 98888-1111", "", "São Paulo", "SP", 5, ""],
]


def _parse_valor(s: str) -> float:
    """Converte string de valor monetário (pt-BR ou en-US) para float.

    Exemplos suportados:
      38.5   → 38.5    (ponto decimal inglês)
      38,5   → 38.5    (vírgula decimal)
      1.234,56 → 1234.56 (pt-BR com milhar)
      1,234.56 → 1234.56 (en-US com milhar)
    """
    s = s.replace("R$", "").strip()
    if "." in s and "," in s:
        # Determine which is the thousands separator
        dot_pos = s.rfind(".")
        comma_pos = s.rfind(",")
        if dot_pos > comma_pos:
            # en-US: 1,234.56
            s = s.replace(",", "")
        else:
            # pt-BR: 1.234,56
            s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    return float(s)


def _azul():
    return PatternFill("solid", fgColor="1A56DB")


def _amarelo():
    return PatternFill("solid", fgColor="FFF8DC")


def _header_font(cor="FFFFFF"):
    return Font(bold=True, color=cor, size=10)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar_template_fornecedores() -> bytes:
    """Gera template XLSX para importação de fornecedores."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fornecedores"

    # Instrução (linha 1)
    ws.append(
        ["INSTRUÇÕES: Preencha a partir da linha 4. "
         "Linhas que começam com 'EXEMPLO' serão ignoradas na importação."]
    )
    ws.merge_cells(f"A1:{get_column_letter(len(_FORN_COLS))}1")
    inst_cell = ws["A1"]
    inst_cell.font = Font(italic=True, color="666666", size=9)
    inst_cell.fill = PatternFill("solid", fgColor="F1F5F9")

    # Cabeçalho (linha 2)
    ws.append([c[0] for c in _FORN_COLS])
    for col_idx in range(1, len(_FORN_COLS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = _azul()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Exemplos (linhas 3–4)
    ex_fill = _amarelo()
    ex_font = Font(color="8B6914", italic=True)
    for ex in _FORN_EXEMPLOS:
        ws.append(ex)
        for col_idx in range(1, len(_FORN_COLS) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = ex_fill
            cell.font = ex_font

    _set_col_widths(ws, [c[1] for c in _FORN_COLS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_fornecedores(fornecedores: list[dict]) -> bytes:
    """Exporta lista de fornecedores para XLSX."""
    _COLS: list[tuple[str, str, int]] = [
        ("Nome / Razão Social", "nome", 35),
        ("CNPJ", "cnpj", 20),
        ("Categoria", "categoria", 15),
        ("Contato", "contato", 20),
        ("Telefone", "telefone", 18),
        ("E-mail", "email", 30),
        ("Cidade", "cidade", 18),
        ("UF", "uf", 5),
        ("Avaliação", "avaliacao", 12),
        ("Ativo", "ativo", 8),
        ("Observações", "observacoes", 40),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Fornecedores"

    ws.append([c[0] for c in _COLS])
    for col_idx in range(1, len(_COLS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _azul()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20
    _set_col_widths(ws, [c[2] for c in _COLS])

    for f in fornecedores:
        ws.append([
            f.get("nome", ""),
            f.get("cnpj") or "",
            f.get("categoria") or "",
            f.get("contato") or "",
            f.get("telefone") or "",
            f.get("email") or "",
            f.get("cidade") or "",
            f.get("uf") or "",
            f.get("avaliacao"),
            "Sim" if f.get("ativo") else "Não",
            f.get("observacoes") or "",
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_fornecedores_xlsx(conteudo: bytes) -> tuple[list[dict], list[str]]:
    """
    Parseia um XLSX de fornecedores.
    Retorna (lista_de_dicts, lista_de_erros).
    Pula: linha de instrução, cabeçalho, linhas 'EXEMPLO' e linhas vazias.
    """
    wb = load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    fornecedores: list[dict] = []
    erros: list[str] = []

    # Encontra linha de cabeçalho (procura "nome" ou "razão" na 1ª célula)
    header_row: int | None = None
    for row in ws.iter_rows(max_row=10):
        first = str(row[0].value or "").strip().lower()
        if "nome" in first or "razão" in first:
            header_row = row[0].row
            break

    if header_row is None:
        return [], ["Cabeçalho não encontrado. Use o template fornecido."]

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = [str(v).strip() if v is not None else "" for v in (row or [])[:10]]
        # Garante tamanho mínimo
        while len(values) < 10:
            values.append("")

        nome = values[0]
        if not nome or nome.upper().startswith("EXEMPLO") or nome.startswith("↑"):
            continue

        avaliacao: float | None = None
        av_str = values[8]
        if av_str:
            try:
                av = float(av_str.replace(",", "."))
                if 1.0 <= av <= 5.0:
                    avaliacao = av
                else:
                    erros.append(
                        f"Linha {row_idx}: Avaliação '{av_str}' fora do intervalo 1–5 (campo ignorado)"
                    )
            except ValueError:
                erros.append(
                    f"Linha {row_idx}: Avaliação inválida '{av_str}' (campo ignorado)"
                )

        uf = values[7].upper()[:2] if values[7] else None

        fornecedores.append({
            "nome": nome,
            "cnpj": values[1] or None,
            "categoria": values[2] or None,
            "contato": values[3] or None,
            "telefone": values[4] or None,
            "email": values[5] or None,
            "cidade": values[6] or None,
            "uf": uf,
            "avaliacao": avaliacao,
            "observacoes": values[9] or None,
            "ativo": True,
        })

    return fornecedores, erros


# ── Requisição — itens ────────────────────────────────────────────────────────

_REQ_ITEM_COLS: list[tuple[str, int]] = [
    ("Descrição do Material *", 42),
    ("Unidade *", 12),
    ("Quantidade *", 14),
    ("Observação", 40),
]

_REQ_ITEM_EXEMPLOS = [
    ["EXEMPLO: Cimento CP-III", "sc", 50, "Entregar no almoxarifado"],
    ["EXEMPLO: Areia lavada fina", "m³", 10, ""],
    ["EXEMPLO: Brita 1", "t", 15, ""],
]


def gerar_template_requisicao() -> bytes:
    """Gera template XLSX para importação de itens de requisição."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens da Requisição"

    ws.append(
        ["INSTRUÇÕES: Preencha a partir da linha 3. "
         "Linhas que começam com 'EXEMPLO' serão ignoradas na importação."]
    )
    ws.merge_cells(f"A1:{get_column_letter(len(_REQ_ITEM_COLS))}1")
    ws["A1"].font = Font(italic=True, color="666666", size=9)
    ws["A1"].fill = PatternFill("solid", fgColor="F1F5F9")

    ws.append([c[0] for c in _REQ_ITEM_COLS])
    for col_idx in range(1, len(_REQ_ITEM_COLS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = _azul()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ex_fill = _amarelo()
    ex_font = Font(color="8B6914", italic=True)
    for ex in _REQ_ITEM_EXEMPLOS:
        ws.append(ex)
        for col_idx in range(1, len(_REQ_ITEM_COLS) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = ex_fill
            cell.font = ex_font

    _set_col_widths(ws, [c[1] for c in _REQ_ITEM_COLS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_requisicao_xlsx(conteudo: bytes) -> tuple[list[dict], list[str]]:
    """Parseia XLSX de itens de requisição. Retorna (itens, erros)."""
    wb = load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    itens: list[dict] = []
    erros: list[str] = []

    header_row: int | None = None
    for row in ws.iter_rows(max_row=10):
        first = str(row[0].value or "").strip().lower()
        if "descri" in first:
            header_row = row[0].row
            break

    if header_row is None:
        return [], ["Cabeçalho não encontrado. Use o template fornecido."]

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = [str(v).strip() if v is not None else "" for v in (row or [])[:4]]
        while len(values) < 4:
            values.append("")

        descricao = values[0]
        if not descricao or descricao.upper().startswith("EXEMPLO"):
            continue

        unidade = values[1] or "un"

        try:
            quantidade = float(values[2].replace(",", ".")) if values[2] else None
        except ValueError:
            erros.append(f"Linha {row_idx}: Quantidade inválida '{values[2]}' — linha ignorada")
            continue

        if not quantidade or quantidade <= 0:
            erros.append(f"Linha {row_idx}: Quantidade deve ser maior que zero — linha ignorada")
            continue

        itens.append({
            "descricao": descricao,
            "unidade": unidade,
            "quantidade": quantidade,
            "observacao": values[3] or None,
        })

    return itens, erros


# ── Ordem de Compra — itens ───────────────────────────────────────────────────

_OC_ITEM_COLS: list[tuple[str, int]] = [
    ("Descrição do Item *", 42),
    ("Unidade *", 12),
    ("Quantidade *", 14),
    ("Preço Unitário *", 16),
    ("Observação", 35),
]

_OC_ITEM_EXEMPLOS = [
    ["EXEMPLO: Cimento CP-III", "sc", 100, 38.50, ""],
    ["EXEMPLO: Areia lavada fina", "m³", 20, 85.00, "Entrega no canteiro"],
    ["EXEMPLO: Brita 1", "t", 30, 65.00, ""],
]


def gerar_template_oc_itens() -> bytes:
    """Gera template XLSX para importação de itens de Ordem de Compra."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens da OC"

    ws.append(
        ["INSTRUÇÕES: Preencha a partir da linha 3. "
         "Linhas que começam com 'EXEMPLO' serão ignoradas na importação."]
    )
    ws.merge_cells(f"A1:{get_column_letter(len(_OC_ITEM_COLS))}1")
    ws["A1"].font = Font(italic=True, color="666666", size=9)
    ws["A1"].fill = PatternFill("solid", fgColor="F1F5F9")

    ws.append([c[0] for c in _OC_ITEM_COLS])
    for col_idx in range(1, len(_OC_ITEM_COLS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = _azul()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ex_fill = _amarelo()
    ex_font = Font(color="8B6914", italic=True)
    for ex in _OC_ITEM_EXEMPLOS:
        ws.append(ex)
        for col_idx in range(1, len(_OC_ITEM_COLS) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = ex_fill
            cell.font = ex_font

    _set_col_widths(ws, [c[1] for c in _OC_ITEM_COLS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_oc_itens_xlsx(conteudo: bytes) -> tuple[list[dict], list[str]]:
    """Parseia XLSX de itens de OC. Retorna (itens, erros)."""
    wb = load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    itens: list[dict] = []
    erros: list[str] = []

    header_row: int | None = None
    for row in ws.iter_rows(max_row=10):
        first = str(row[0].value or "").strip().lower()
        if "descri" in first:
            header_row = row[0].row
            break

    if header_row is None:
        return [], ["Cabeçalho não encontrado. Use o template fornecido."]

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = [str(v).strip() if v is not None else "" for v in (row or [])[:5]]
        while len(values) < 5:
            values.append("")

        descricao = values[0]
        if not descricao or descricao.upper().startswith("EXEMPLO"):
            continue

        unidade = values[1] or "un"

        try:
            quantidade = float(values[2].replace(",", ".")) if values[2] else None
        except ValueError:
            erros.append(f"Linha {row_idx}: Quantidade inválida '{values[2]}' — linha ignorada")
            continue

        try:
            preco = _parse_valor(values[3]) if values[3] else None
        except ValueError:
            erros.append(f"Linha {row_idx}: Preço inválido '{values[3]}' — linha ignorada")
            continue

        if not quantidade or quantidade <= 0:
            erros.append(f"Linha {row_idx}: Quantidade deve ser maior que zero — linha ignorada")
            continue
        if preco is None or preco < 0:
            erros.append(f"Linha {row_idx}: Preço unitário inválido — linha ignorada")
            continue

        itens.append({
            "descricao": descricao,
            "unidade": unidade,
            "quantidade": quantidade,
            "preco_unitario": preco,
            "observacao": values[4] or None,
        })

    return itens, erros
